"""
ULB Audit Framework - Report Generator (with Timestamped Folders)
- Creates timestamped subfolder for each audit run
- Prevents report flooding in base folder
"""

import logging
import math
import re
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

class ReportGenerator:
    """Generates audit reports in timestamped folders"""
    
    def __init__(self, output_folder):
        # Create timestamped subfolder
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        self.output_folder = Path(output_folder) / f"Audit_{timestamp}"
        self.output_folder.mkdir(parents=True, exist_ok=True)
        self.logger = logging.getLogger(__name__)
        self.logger.info(f"Reports will be saved to: {self.output_folder}")
        self.column_map = self._load_column_map()

    def _load_column_map(self):
        """Load optional column label/section mapping from mp-map.xlsx."""
        map_path = Path(__file__).resolve().parents[2] / 'mp-map.xlsx'
        if not map_path.exists():
            self.logger.warning("mp-map.xlsx not found. Reports will show raw column names only.")
            return {}

        try:
            df = pd.read_excel(map_path, sheet_name='consolidated', header=None)
            mapping = {}
            for _, row in df.iterrows():
                col_key = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ''
                if not col_key:
                    continue
                section = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ''
                label = str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) else ''
                mapping[col_key] = {
                    'section': section,
                    'label': label,
                }
            self.logger.info(f"[OK] Loaded {len(mapping)} column mappings from mp-map.xlsx")
            return mapping
        except Exception as e:
            self.logger.warning(f"Unable to load mp-map.xlsx: {str(e)}")
            return {}

    def _get_field_label(self, column_code: str) -> str:
        if pd.isna(column_code) or str(column_code).strip() == '':
            return ''
        column_code = str(column_code).strip()
        map_item = self.column_map.get(column_code, {})
        return map_item.get('label') or column_code

    def _get_field_section(self, column_code: str) -> str:
        if pd.isna(column_code) or str(column_code).strip() == '':
            return ''
        column_code = str(column_code).strip()
        map_item = self.column_map.get(column_code, {})
        return map_item.get('section') or ''

    def _format_field_list(self, column_spec: str) -> str:
        if pd.isna(column_spec) or str(column_spec).strip() == '':
            return ''
        columns = [c.strip() for c in str(column_spec).split(',')]
        labels = [self._get_field_label(col) for col in columns if col]
        return ", ".join(labels)

    def _format_section_list(self, column_spec: str) -> str:
        if pd.isna(column_spec) or str(column_spec).strip() == '':
            return ''
        columns = [c.strip() for c in str(column_spec).split(',')]
        sections = []
        for col in columns:
            section = self._get_field_section(col)
            if section and section not in sections:
                sections.append(section)
        return "; ".join(sections)

    def _format_number(self, value):
        if value is None:
            return ''
        try:
            numeric = float(value)
        except (TypeError, ValueError):
            return str(value)
        if math.isnan(numeric):
            return ''
        abs_val = abs(numeric)
        if abs_val >= 100:
            return f"{numeric:,.0f}"
        if abs_val >= 10:
            return f"{numeric:,.1f}".rstrip('0').rstrip('.')
        return f"{numeric:,.2f}".rstrip('0').rstrip('.')

    def _format_operator_phrase(self, operator: str) -> str:
        return {
            '>=': 'at least',
            '<=': 'at most',
            '>': 'greater than',
            '<': 'less than',
            '==': 'equal to',
            '=': 'equal to',
            '!=': 'not equal to',
        }.get(operator, operator)

    def _build_finding_components(self, finding: dict) -> dict:
        check_label = str(finding.get('description', '')).strip()
        detail = str(finding.get('detail', '')).strip()
        check_type = str(finding.get('check_type', '')).lower().strip()

        finding_text = "This check did not meet the expected condition."
        why_text = "This value is outside the acceptable range defined for this data item."

        if detail.startswith('Unable to evaluate:'):
            reason = detail.replace('Unable to evaluate:', '').strip()
            finding_text = f"The check could not be evaluated because {reason}."
            why_text = "This check could not be evaluated due to missing or invalid data."
        elif check_type in {'outlier_iqr', 'outlier_zscore'}:
            match = re.search(
                r'Value\s+(?P<value>-?\d+(\.\d+)?)\s+is\s+'
                r'(?P<position>below lower bound|above upper bound)\s+'
                r'(?P<bound>-?\d+(\.\d+)?)',
                detail
            )
            if match:
                value = self._format_number(match.group('value'))
                bound = self._format_number(match.group('bound'))
                position = match.group('position')
                if 'below' in position:
                    finding_text = (
                        f"The calculated value is {value}, which is lower than the expected lower bound of {bound}."
                    )
                else:
                    finding_text = (
                        f"The calculated value is {value}, which is higher than the expected upper bound of {bound}."
                    )
            else:
                finding_text = "This value appears statistically unusual compared with similar municipalities."
            why_text = (
                "This value is significantly higher/lower than most similar municipalities and appears "
                "statistically unusual."
            )
        elif detail.startswith('Missing/zero:'):
            missing_cols = detail.replace('Missing/zero:', '').strip()
            cols = [c.strip() for c in missing_cols.split(',') if c.strip()]
            labels = [self._get_field_label(col) for col in cols]
            finding_text = (
                "The following required fields are missing or zero: "
                f"{', '.join(labels) if labels else missing_cols}."
            )
            why_text = "One or more required fields were missing or zero for this check."
        elif detail.startswith('Cross-table mismatch:'):
            match = re.search(
                r"primary column '(.+?)'\s*=\s*(?P<primary>-?\d+(\.\d+)?),\s*"
                r"reference column '(.+?)'\s*=\s*(?P<reference>-?\d+(\.\d+)?)",
                detail
            )
            if match:
                primary_value = self._format_number(match.group('primary'))
                reference_value = self._format_number(match.group('reference'))
                finding_text = (
                    f"The primary field value is {primary_value}, while the reference field value is "
                    f"{reference_value}, and they do not match."
                )
            else:
                finding_text = "The primary field value does not match the reference field value."
            why_text = "Values from related tables should be consistent with each other."
        elif detail.startswith('Consistency:'):
            comp = detail.replace('Consistency:', '').strip()
            match = re.search(
                r'(?P<value>-?\d+(\.\d+)?)\s*(?P<op>!=|==|>=|<=|>|<)\s*(?P<other>-?\d+(\.\d+)?)',
                comp
            )
            if match:
                value = self._format_number(match.group('value'))
                other = self._format_number(match.group('other'))
                operator_phrase = self._format_operator_phrase(match.group('op'))
                finding_text = (
                    f"The value is {value}, which does not meet the expected condition of being {operator_phrase} "
                    f"{other}."
                )
            else:
                finding_text = "The values did not meet the expected consistency condition."
            why_text = "Values that should be consistent were found to differ."
        else:
            range_match = re.search(
                r'(?P<value>-?\d+(\.\d+)?)\s+not in range\s+\[(?P<lower>-?\d+(\.\d+)?),\s*'
                r'(?P<upper>-?\d+(\.\d+)?)\]',
                detail
            )
            comp_match = re.search(
                r'(?P<value>-?\d+(\.\d+)?)\s+not\s+(?P<op>!=|==|>=|<=|>|<)\s+'
                r'(?P<threshold>-?\d+(\.\d+)?)',
                detail
            )
            if range_match:
                value = self._format_number(range_match.group('value'))
                lower = self._format_number(range_match.group('lower'))
                upper = self._format_number(range_match.group('upper'))
                finding_text = (
                    f"The calculated value is {value}, which is outside the expected range of {lower} to {upper}."
                )
            elif comp_match:
                value = self._format_number(comp_match.group('value'))
                threshold = self._format_number(comp_match.group('threshold'))
                operator_phrase = self._format_operator_phrase(comp_match.group('op'))
                finding_text = (
                    f"The value is {value}, which does not meet the expected condition of being {operator_phrase} "
                    f"{threshold}."
                )
            elif detail:
                finding_text = detail
            finding_text = f"{finding_text} This suggests the input values for this check may need review."

        primary_field = self._format_field_list(finding.get('column_1'))
        reference_field = self._format_field_list(finding.get('column_2'))
        section = self._format_section_list(finding.get('column_1'))

        fields_lines = []
        if primary_field:
            fields_lines.append(f"Primary Field: {primary_field}")
        if reference_field:
            fields_lines.append(f"Reference Field: {reference_field}")
        if section:
            fields_lines.append(f"Section: {section}")

        return {
            'check_label': check_label,
            'finding_text': finding_text,
            'why_text': why_text,
            'fields_lines': fields_lines,
        }
        
    def generate_ulb_report(self, mp_id, ulb_info, findings):
        """Generate individual ULB audit report PDF"""
        filename = f"Audit_Report_{ulb_info['municipality_name'].replace(' ', '_')}_ID{mp_id}.pdf"
        filepath = self.output_folder / filename
        
        doc = SimpleDocTemplate(str(filepath), pagesize=A4,
                               topMargin=0.5*inch, bottomMargin=0.5*inch,
                               leftMargin=0.5*inch, rightMargin=0.5*inch)
        
        story = []
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#1f4788'),
            spaceAfter=12,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=12,
            textColor=colors.HexColor('#2e5c8a'),
            spaceAfter=6,
            spaceBefore=12,
            fontName='Helvetica-Bold'
        )
        
        story.append(Paragraph("TAMIL NADU STATE FINANCE COMMISSION", title_style))
        story.append(Paragraph("ULB Questionnaire Audit Report", title_style))
        story.append(Spacer(1, 0.2*inch))
        
        ulb_info_data = [
            ["Municipality", ulb_info['municipality_name']],
            ["District", ulb_info['district_name']],
            ["ULB ID", str(mp_id)],
            ["Report Date", datetime.now().strftime("%d-%b-%Y")],
        ]
        
        ulb_table = Table(ulb_info_data, colWidths=[2*inch, 4*inch])
        ulb_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e8f0f8')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        story.append(ulb_table)
        story.append(Spacer(1, 0.3*inch))
        
        story.append(Paragraph("EXECUTIVE SUMMARY", heading_style))
        
        if not findings:
            story.append(Paragraph("No audit observations found. All validations passed successfully.", 
                                 styles['Normal']))
        else:
            severity_counts = {}
            for f in findings:
                sev = f['severity']
                severity_counts[sev] = severity_counts.get(sev, 0) + 1
            
            summary_text = f"Total Observations: {len(findings)}<br/>"
            for sev in ['Critical', 'High', 'Medium', 'Low']:
                if sev in severity_counts:
                    summary_text += f"• {sev} Priority: {severity_counts[sev]}<br/>"
            
            story.append(Paragraph(summary_text, styles['Normal']))
        
        story.append(Spacer(1, 0.3*inch))
        
        if findings:
            story.append(Paragraph("DETAILED OBSERVATIONS", heading_style))
            
            findings_by_part = {}
            for f in findings:
                part = f['part_no']
                if part not in findings_by_part:
                    findings_by_part[part] = []
                findings_by_part[part].append(f)
            
            def part_sort_key(part_str):
                try:
                    return int(str(part_str).split(',')[0])
                except:
                    return 999
            
            for part in sorted(findings_by_part.keys(), key=part_sort_key):
                part_findings = findings_by_part[part]
                
                part_names = {
                    '1': 'PART I - DEMOGRAPHICS AND GEOGRAPHY',
                    '2': 'PART II - HUMAN RESOURCES',
                    '3': 'PART III - ACCOUNTS',
                    '4': 'PART IV - TAXATION AND DCB',
                    '5': 'PART V - LIABILITIES',
                    '6': 'PART VI - CAPITAL WORKS',
                    '7': 'PART VII - ASSETS',
                    '8': 'PART VIII - SERVICE LEVELS',
                    '9': 'PART IX - FUTURE NEEDS',
                }
                
                part_name = part_names.get(str(part), f'PART {part}')
                story.append(Paragraph(part_name, heading_style))
                
                table_data = [['#', 'Rule', 'Severity', 'Observation']]
                
                for idx, f in enumerate(part_findings, 1):
                    components = self._build_finding_components(f)
                    fields_block = "<br/>".join(components['fields_lines']) if components['fields_lines'] else "N/A"
                    obs_text = (
                        f"<b>Check:</b> {components['check_label']}<br/>"
                        f"<b>Finding:</b><br/>{components['finding_text']}<br/>"
                        f"<b>Why this was flagged:</b><br/>{components['why_text']}<br/>"
                        f"<b>Fields Involved:</b><br/>{fields_block}"
                    )
                    table_data.append([
                        str(idx),
                        f['rule_id'],
                        f['severity'],
                        Paragraph(obs_text, styles['Normal'])
                    ])
                
                findings_table = Table(table_data, colWidths=[0.3*inch, 0.8*inch, 0.8*inch, 5*inch])
                
                table_style = [
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                ]
                
                for idx in range(1, len(table_data)):
                    severity = table_data[idx][2]
                    severity_color = {
                        'Critical': colors.HexColor('#ff0000'),
                        'High': colors.HexColor('#ffc000'),
                        'Medium': colors.HexColor('#ffff00'),
                        'Low': colors.HexColor('#92d050')
                    }.get(severity, colors.white)
                    table_style.append(('BACKGROUND', (2, idx), (2, idx), severity_color))
                
                findings_table.setStyle(TableStyle(table_style))
                story.append(findings_table)
                story.append(Spacer(1, 0.2*inch))
        
        story.append(Spacer(1, 0.3*inch))
        footer_text = """
        <para align=justify>
        <b>Note:</b> This audit report is generated automatically based on validation rules applied to the 
        ULB questionnaire data. All findings should be reviewed and verified.
        </para>
        """
        story.append(Paragraph(footer_text, styles['Normal']))
        
        doc.build(story)
        self.logger.info(f"[OK] Generated report: {filename}")
        return filepath

    def generate_ulb_excel_report(self, mp_id, ulb_info, findings):
        """Generate individual ULB audit report Excel with remark entry."""
        filename = f"Audit_Report_{ulb_info['municipality_name'].replace(' ', '_')}_ID{mp_id}.xlsx"
        filepath = self.output_folder / filename

        wb = Workbook()
        ws = wb.active
        ws.title = "Audit Findings"

        title_font = Font(size=14, bold=True, color="1F4788")
        heading_font = Font(size=11, bold=True, color="2E5C8A")
        header_font = Font(size=10, bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="366092")
        wrap_alignment = Alignment(wrap_text=True, vertical="top")
        center_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin", color="999999"),
            right=Side(style="thin", color="999999"),
            top=Side(style="thin", color="999999"),
            bottom=Side(style="thin", color="999999"),
        )

        ws.merge_cells("A1:J1")
        ws["A1"] = "TAMIL NADU STATE FINANCE COMMISSION - ULB Questionnaire Audit Report"
        ws["A1"].font = title_font
        ws["A1"].alignment = center_alignment

        ws["A3"] = "Municipality"
        ws["B3"] = ulb_info["municipality_name"]
        ws["A4"] = "District"
        ws["B4"] = ulb_info["district_name"]
        ws["A5"] = "ULB ID"
        ws["B5"] = str(mp_id)
        ws["A6"] = "Report Date"
        ws["B6"] = datetime.now().strftime("%d-%b-%Y")

        for row in range(3, 7):
            ws[f"A{row}"].font = heading_font

        start_row = 8
        headers = [
            "#",
            "Rule",
            "Severity",
            "Check",
            "Finding",
            "Why this was flagged",
            "Primary Field",
            "Reference Field",
            "Section",
            "Remarks",
        ]
        ws.append([""] * len(headers))
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border

        row_cursor = start_row + 1
        if findings:
            for idx, f in enumerate(findings, 1):
                components = self._build_finding_components(f)
                fields_lines = components['fields_lines']
                primary_field = next((line.replace("Primary Field: ", "") for line in fields_lines if line.startswith("Primary Field: ")), "")
                reference_field = next((line.replace("Reference Field: ", "") for line in fields_lines if line.startswith("Reference Field: ")), "")
                section = next((line.replace("Section: ", "") for line in fields_lines if line.startswith("Section: ")), "")

                row_values = [
                    idx,
                    f.get("rule_id", ""),
                    f.get("severity", ""),
                    components["check_label"],
                    components["finding_text"],
                    components["why_text"],
                    primary_field,
                    reference_field,
                    section,
                    "",
                ]
                for col, value in enumerate(row_values, 1):
                    cell = ws.cell(row=row_cursor, column=col, value=value)
                    cell.alignment = wrap_alignment if col >= 4 else center_alignment
                    cell.border = thin_border
                row_cursor += 1
        else:
            ws.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=10)
            cell = ws.cell(row=row_cursor, column=1, value="No audit observations found. All validations passed successfully.")
            cell.alignment = wrap_alignment
            row_cursor += 1

        column_widths = [5, 10, 10, 28, 45, 40, 30, 30, 30, 25]
        for idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = width

        ws.freeze_panes = "A9"
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.print_area = f"A1:J{row_cursor - 1}"

        ws.protection.sheet = True
        ws.protection.enable()
        for row in ws.iter_rows(min_row=start_row + 1, max_row=row_cursor - 1, min_col=1, max_col=10):
            for cell in row:
                cell.protection = cell.protection.copy(locked=True)
        for row in ws.iter_rows(min_row=start_row + 1, max_row=row_cursor - 1, min_col=10, max_col=10):
            for cell in row:
                cell.protection = cell.protection.copy(locked=False)

        wb.save(filepath)
        self.logger.info(f"[OK] Generated Excel report: {filename}")
        return filepath
    
    def generate_master_dashboard(self, all_findings, data_loader):
        """Generate master dashboard Excel"""
        filepath = self.output_folder / f"Master_Audit_Dashboard.xlsx"
        
        with pd.ExcelWriter(str(filepath), engine='openpyxl') as writer:
            summary_data = []
            for ulb in data_loader.ulb_list:
                mp_id = ulb['mp_id']
                ulb_findings = [f for f in all_findings if f['mp_id'] == mp_id]
                
                severity_counts = {'Critical': 0, 'High': 0, 'Medium': 0, 'Low': 0}
                for f in ulb_findings:
                    sev = f['severity']
                    if sev in severity_counts:
                        severity_counts[sev] += 1
                
                summary_data.append({
                    'ULB_ID': mp_id,
                    'Municipality': ulb['municipality_name'],
                    'District': ulb['district_name'],
                    'Total_Findings': len(ulb_findings),
                    'Critical': severity_counts['Critical'],
                    'High': severity_counts['High'],
                    'Medium': severity_counts['Medium'],
                    'Low': severity_counts['Low'],
                })
            
            df_summary = pd.DataFrame(summary_data)
            df_summary.to_excel(writer, sheet_name='Summary', index=False)
            
            if all_findings:
                df_findings = pd.DataFrame(all_findings)
                df_findings = df_findings[[
                    'mp_id', 'ulb_name', 'district', 'rule_id', 'part_no', 
                    'severity', 'check_type', 'description', 'detail'
                ]]
                df_findings.columns = [
                    'ULB_ID', 'Municipality', 'District', 'Rule_ID', 'Part', 
                    'Severity', 'Check_Type', 'Validation_Rule', 'Finding_Detail'
                ]
                df_findings.to_excel(writer, sheet_name='All_Findings', index=False)
                
                for sev in ['Critical', 'High']:
                    df_sev = df_findings[df_findings['Severity'] == sev]
                    if not df_sev.empty:
                        sheet_name = f'{sev}_Priority'
                        df_sev.to_excel(writer, sheet_name=sheet_name, index=False)
            else:
                df_empty = pd.DataFrame({'Message': ['No audit findings. All validations passed.']})
                df_empty.to_excel(writer, sheet_name='All_Findings', index=False)
        
        self.logger.info(f"[OK] Generated master dashboard: {filepath.name}")
        return filepath
    
    def generate_master_tabular_report(self, all_findings, data_loader):
        """Generate detailed tabular master report Excel"""
        filepath = self.output_folder / f"Master_Detailed_Report.xlsx"
        
        with pd.ExcelWriter(str(filepath), engine='openpyxl') as writer:
            if all_findings:
                df_all = pd.DataFrame(all_findings)
                
                def part_sort_key(part_str):
                    try:
                        return int(str(part_str))
                    except:
                        return 999
                
                for part_no in sorted(df_all['part_no'].unique(), key=part_sort_key):
                    df_part = df_all[df_all['part_no'] == part_no]
                    
                    df_output = df_part[[
                        'mp_id', 'ulb_name', 'district', 'rule_id', 
                        'severity', 'description', 'detail'
                    ]].copy()
                    
                    df_output.columns = [
                        'ULB_ID', 'Municipality', 'District', 'Rule_ID',
                        'Severity', 'Validation_Rule', 'Finding_Detail'
                    ]
                    
                    sheet_name = f'Part_{part_no}'[:31]
                    df_output.to_excel(writer, sheet_name=sheet_name, index=False)
            else:
                df_empty = pd.DataFrame({'Message': ['No audit findings. All validations passed.']})
                df_empty.to_excel(writer, sheet_name='Summary', index=False)
        
        self.logger.info(f"[OK] Generated detailed report: {filepath.name}")
        return filepath
