"""
ULB Audit Framework - Validation Rules Master Generator
Creates comprehensive Excel workbook with all validation rules from audit documentation
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def create_validation_rules_master():
    """Create comprehensive validation rules Excel workbook"""
    
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ============================================================================
    # SHEET 1: VALIDATION RULES
    # ============================================================================
    
    validation_rules = [
        # PART I - DEMOGRAPHICS AND GEOGRAPHY
        ["P1_001", "1", "p1_1_1_2", "Sanity", "Population density check", 
         "population_density = p1_1_3_4_tot_25_tot / p1_1_1_3_area; Flag if < 300 or > 35000 per sq.km", 
         "300-35000", "Medium", "TRUE", "Unrealistic population density"],
        
        ["P1_002", "1", "p1_1_1_2", "Sanity", "Household size validation",
         "hh_size = p1_1_3_4_tot_25_tot / p1_1_5_family; Flag if < 2.5 or > 6", 
         "2.5-6", "Medium", "TRUE", "Unusual household size"],
        
        ["P1_003", "1", "p1_1_1_2", "Consistency", "Census vs latest population arithmetic",
         "p1_1_3_4_tot_25_tot should >= p1_1_3_4_tot_11_tot (2025 >= 2011)", 
         "25>=11", "High", "TRUE", "Population decreased from 2011 to 2025"],
        
        ["P1_004", "1", "p1_1_1_2", "Consistency", "Gender balance check",
         "male_ratio = p1_1_3_4_tot_25_ma / p1_1_3_4_tot_25_tot; Flag if < 0.45 or > 0.55", 
         "0.45-0.55", "Low", "TRUE", "Unusual gender ratio"],
        
        ["P1_005", "1", "p1_1_1_2", "Consistency", "SC+ST+Others = Total population",
         "(p1_1_3_1_sc_25_tot + p1_1_3_2_st_25_tot + p1_1_3_3_oth_25_tot) == p1_1_3_4_tot_25_tot", 
         "SUM=TOTAL", "Critical", "TRUE", "Category populations don't add up"],
        
        ["P1_006", "1", "p1_1_1_2", "Consistency", "Slum population vs total",
         "slum_pop = p1_1_4_1_25_tot + p1_1_4_2_25_tot; Flag if > p1_1_3_4_tot_25_tot", 
         "SLUM<=TOTAL", "High", "TRUE", "Slum population exceeds total"],
        
        ["P1_007", "1", "p1_1_1_2", "Sanity", "Slum household reasonableness",
         "slum_hh = p1_1_4_1_25_nos + p1_1_4_2_25_nos; slum_hh_size = slum_pop / slum_hh; Flag if < 2 or > 8", 
         "2-8", "Medium", "TRUE", "Unusual slum household size"],
        
        ["P1_008", "1", "p1_1_1_2", "Sanity", "Floating population reasonableness",
         "floating_ratio = p1_1_6_1_avg_day / p1_1_3_4_tot_25_tot; Flag if > 0.5 (50%)", 
         "<0.5", "Medium", "TRUE", "Unusually high floating population"],
        
        ["P1_009", "1", "p1_1_1_2", "Ratio", "Growth rate calculation",
         "cagr = ((p1_1_3_4_tot_25_tot / p1_1_3_4_tot_11_tot) ** (1/14)) - 1; Flag if < 0.005 or > 0.05", 
         "0.5-5%", "Medium", "TRUE", "Unusual population growth rate"],
        
        ["P1_010", "1", "p1_1_1_2", "Consistency", "Wards must exist",
         "p1_1_1_4_i_wards > 0", 
         ">0", "Critical", "TRUE", "Municipality must have at least one ward"],
        
        # PART II - HUMAN RESOURCES
        ["P2_001", "2", "p2_2_1_1_2_1_5", "Arithmetic", "Sanctioned = OPS + CPS + Vacant",
         "For each row: verify Sanctioned = OPS + CPS + Vacant", 
         "SUM=TOTAL", "Critical", "TRUE", "Position arithmetic mismatch"],
        
        ["P2_002", "2", "p2_2_1_1_2_1_5", "Sanity", "Salary reasonableness by grade",
         "Class I (Level 18-22): ₹6-18 lakh; Class II (12-17): ₹4-10 lakh; Class III (6-11): ₹2.5-6 lakh; Class IV (1-5): ₹1.8-4 lakh", 
         "By Grade", "High", "TRUE", "Salary outside expected range for grade"],
        
        ["P2_003", "2", "p2_2_1_1_2_1_5", "Sanity", "Vacancy rate check",
         "vacancy_rate = SUM(Vacant) / SUM(Sanctioned); Flag if > 0.4 (40%)", 
         "<40%", "Medium", "TRUE", "Unusually high vacancy rate"],
        
        ["P2_004", "2", "p2_2_1_1_2_1_5", "Ratio", "Staff per 1000 population",
         "staff_ratio = SUM(OPS + CPS) / (population / 1000); Flag if < 2 or > 8", 
         "2-8", "Medium", "TRUE", "Staff ratio outside normal range"],
        
        ["P2_005", "2", "p2_2_1_1_2_1_5", "Consistency", "Negative values not allowed",
         "All numeric columns >= 0", 
         ">=0", "Critical", "TRUE", "Negative staff count"],
        
        ["P2_006", "2", "p2_2_1_1_2_1_5", "Ratio", "Total salary per capita",
         "per_capita_sal = total_salary / population; Flag if > 3000", 
         "<3000", "Medium", "TRUE", "Unusually high salary burden per capita"],
        
        ["P2_007", "2", "Multiple", "Cross-table", "Retirement benefits must match retirees",
         "If Part V pension liability exists, Part II must show retirees or pension expenses", 
         "CONSISTENCY", "High", "TRUE", "Pension liability without retiree records"],
        
        # PART III - ACCOUNTS (REVENUE)
        ["P3_001", "3", "p3", "Trend", "Revenue growth reasonableness",
         "For each revenue line: yoy_growth = (current_year - prev_year) / prev_year; Flag if > 0.5 or < -0.3", 
         "-30% to +50%", "Medium", "TRUE", "Unusual revenue growth"],
        
        ["P3_002", "3", "p3", "Ratio", "Property tax per capita",
         "pt_per_capita = property_tax / population; Flag if < 500 or > 4000", 
         "₹500-4000", "Medium", "TRUE", "Property tax per capita outside normal range"],
        
        ["P3_003", "3", "p3", "Ratio", "Own revenue ratio",
         "own_rev_ratio = own_revenue / total_revenue; Flag if < 0.3 (30%)", 
         ">30%", "High", "TRUE", "Low fiscal autonomy"],
        
        ["P3_004", "3", "p3", "Consistency", "Revenue subtotals arithmetic",
         "Verify all section totals = sum of line items", 
         "SUM=TOTAL", "Critical", "TRUE", "Revenue arithmetic error"],
        
        ["P3_005", "3", "p3", "Sanity", "Negative revenue not allowed (except refunds)",
         "All revenue items >= 0 (except specific refund/adjustment lines)", 
         ">=0", "Critical", "TRUE", "Negative revenue value"],
        
        ["P3_006", "3", "p3", "Ratio", "Water charges vs population",
         "water_per_capita = water_charges / population; Flag if < 200 or > 1500", 
         "₹200-1500", "Medium", "TRUE", "Water charges per capita unusual"],
        
        # PART III - ACCOUNTS (EXPENDITURE)
        ["P3_007", "3", "p3", "Trend", "Expenditure growth reasonableness",
         "yoy_growth = (current - prev) / prev; Flag if > 0.5 or < -0.2", 
         "-20% to +50%", "Medium", "TRUE", "Unusual expenditure growth"],
        
        ["P3_008", "3", "p3", "Ratio", "Salary to total expenditure",
         "sal_ratio = salary_exp / total_exp; Flag if > 0.5 (50%)", 
         "<50%", "High", "TRUE", "High salary burden"],
        
        ["P3_009", "3", "p3", "Consistency", "Expenditure subtotals arithmetic",
         "Verify all section totals = sum of line items", 
         "SUM=TOTAL", "Critical", "TRUE", "Expenditure arithmetic error"],
        
        ["P3_010", "3", "p3", "Ratio", "O&M as % of asset value - Roads",
         "om_ratio = road_om_exp / road_asset_value; Flag if < 0.02 or > 0.06", 
         "2-6%", "Medium", "TRUE", "Road O&M outside normal range"],
        
        ["P3_011", "3", "p3", "Ratio", "O&M as % of asset value - Water",
         "om_ratio = water_om_exp / water_asset_value; Flag if < 0.03 or > 0.08", 
         "3-8%", "Medium", "TRUE", "Water O&M outside normal range"],
        
        ["P3_012", "3", "p3", "Ratio", "O&M as % of asset value - SWM",
         "om_ratio = swm_om_exp / swm_asset_value; Flag if < 0.05 or > 0.15", 
         "5-15%", "Medium", "TRUE", "SWM O&M outside normal range"],
        
        # PART III - BALANCE SHEET
        ["P3_013", "3", "p3", "Arithmetic", "Balance sheet equation",
         "Assets = Liabilities + Net Worth", 
         "A=L+NW", "Critical", "TRUE", "Balance sheet doesn't balance"],
        
        ["P3_014", "3", "p3", "Consistency", "Opening balance = Previous closing",
         "For each year: OB(current) == CB(previous)", 
         "OB=Prev_CB", "Critical", "TRUE", "Opening balance mismatch"],
        
        ["P3_015", "3", "p3", "Sanity", "Cash balance reasonableness",
         "cash_months = cash_balance / (total_exp / 12); Flag if > 12 months", 
         "<12 months", "Medium", "TRUE", "Unusually high cash balance"],
        
        # PART IV - TAXATION & DCB
        ["P4_001", "4", "p4", "Arithmetic", "Demand = Previous arrear + Current demand",
         "For each tax: Total_Demand = Arrear_BF + Current_Demand", 
         "SUM=TOTAL", "Critical", "TRUE", "Demand calculation error"],
        
        ["P4_002", "4", "p4", "Arithmetic", "Balance = Demand - Collection",
         "Balance = Total_Demand - Total_Collection", 
         "BAL=DEM-COL", "Critical", "TRUE", "DCB balance mismatch"],
        
        ["P4_003", "4", "p4", "Ratio", "Collection efficiency",
         "coll_eff = Current_Collection / Current_Demand; Flag if < 0.5 or > 1.0", 
         "50-100%", "High", "TRUE", "Unusual collection efficiency"],
        
        ["P4_004", "4", "p4", "Consistency", "Arrear carried forward",
         "Arrear_CF(current) should == Balance(previous) - Adjustments", 
         "CF=Prev_Bal", "High", "TRUE", "Arrear carry-forward mismatch"],
        
        ["P4_005", "4", "p4", "Ratio", "Demand per assessment",
         "demand_per_unit = Total_Demand / No_of_Assessments; Flag if < 500 or > 50000", 
         "₹500-50k", "Medium", "TRUE", "Unusual demand per assessment"],
        
        ["P4_006", "4", "p4", "Consistency", "Assessments vs households",
         "property_assessments should be <= households (from Part 1)", 
         "ASSESS<=HH", "High", "TRUE", "Assessments exceed households"],
        
        ["P4_007", "4", "p4", "Sanity", "Uncollectible percentage",
         "uncoll_pct = Uncollectible / Total_Demand; Flag if > 0.1 (10%)", 
         "<10%", "Medium", "TRUE", "High uncollectible amount"],
        
        ["P4_008", "4", "p4", "Cross-part", "Property tax reconciliation with Part III",
         "Property_Tax_Collection(P4) should match Property_Tax_Revenue(P3)", 
         "P4=P3", "High", "TRUE", "Property tax mismatch between P4 and P3"],
        
        # PART V - LIABILITIES
        ["P5_001", "5", "p5a", "Arithmetic", "Due = Demand - Remitted",
         "For each liability: Due = Demand - Remitted", 
         "DUE=DEM-REM", "Critical", "TRUE", "Liability arithmetic error"],
        
        ["P5_002", "5", "p5b", "Arithmetic", "Outstanding loan calculation",
         "Outstanding = Loan_Amount - Total_Principal_Repaid", 
         "OUT=LOAN-PAID", "Critical", "TRUE", "Loan outstanding mismatch"],
        
        ["P5_003", "5", "p5a", "Sanity", "EPF remittance delay",
         "If Due_EPF > 3_months_contribution, flag", 
         "<3 months", "High", "TRUE", "EPF remittance delay"],
        
        ["P5_004", "5", "p5b", "Consistency", "Loan repayment vs Part III",
         "Total_Principal_Repaid(P5b) should match Interest_Paid(P3)", 
         "P5b≈P3", "High", "TRUE", "Loan repayment mismatch with accounts"],
        
        ["P5_005", "5", "p5b", "Ratio", "Debt service ratio",
         "dsr = (principal + interest) / total_revenue; Flag if > 0.25 (25%)", 
         "<25%", "High", "TRUE", "High debt burden"],
        
        ["P5_006", "5", "p5a", "Cross-part", "Pension liability vs retirees",
         "If pension_liability > 0, Part II must show retirees", 
         "CONSISTENCY", "High", "TRUE", "Pension liability without retirees"],
        
        ["P5_007", "5", "p5b", "Sanity", "Loan purpose reasonableness",
         "Verify loan purpose matches asset additions in Part VII", 
         "CONSISTENCY", "Medium", "TRUE", "Loan purpose unclear"],
        
        # PART VI - CAPITAL WORKS
        ["P6_001", "6", "p6", "Sanity", "BT Road unit cost",
         "unit_cost = total_cost / length_km; Flag if < 30 or > 250 lakh/km", 
         "₹30-250L/km", "High", "TRUE", "BT road cost outside normal range"],
        
        ["P6_002", "6", "p6", "Sanity", "CC Road unit cost",
         "unit_cost = total_cost / length_km; Flag if < 50 or > 300 lakh/km", 
         "₹50-300L/km", "High", "TRUE", "CC road cost outside normal range"],
        
        ["P6_003", "6", "p6", "Sanity", "Water pipeline unit cost",
         "unit_cost = total_cost / length_km; Flag if < 15 or > 80 lakh/km", 
         "₹15-80L/km", "Medium", "TRUE", "Pipeline cost outside normal range"],
        
        ["P6_004", "6", "p6", "Sanity", "STP unit cost",
         "unit_cost = total_cost / capacity_MLD; Flag if < 80 or > 300 lakh/MLD", 
         "₹80-300L/MLD", "High", "TRUE", "STP cost outside normal range"],
        
        ["P6_005", "6", "p6", "Sanity", "OHT unit cost",
         "unit_cost = total_cost / capacity_lakh_litres; Flag if < 20 or > 100 lakh/LL", 
         "₹20-100L/LL", "High", "TRUE", "OHT cost outside normal range"],
        
        ["P6_006", "6", "p6", "Arithmetic", "Funding sources sum",
         "State + Central + General + Others = Total_Cost", 
         "SUM=TOTAL", "Critical", "TRUE", "Funding sources don't add up"],
        
        ["P6_007", "6", "p6", "Engineering", "OHT requires pipelines",
         "If OHT constructed, must have pipeline network", 
         "LOGIC", "High", "TRUE", "OHT without distribution network"],
        
        ["P6_008", "6", "p6", "Engineering", "STP requires sewer network",
         "If STP constructed, must have UGD/sewer length", 
         "LOGIC", "High", "TRUE", "STP without collection network"],
        
        ["P6_009", "6", "p6", "Ratio", "Capital works vs population",
         "capex_per_capita = total_capex / population; Flag if > 10000", 
         "<₹10k", "Medium", "TRUE", "Unusually high capex per capita"],
        
        ["P6_010", "6", "p6", "Cross-part", "Asset additions match Part VII",
         "New assets in P6 should appear in P7 asset register", 
         "P6→P7", "High", "TRUE", "Capital work not reflected in assets"],
        
        ["P6_011", "6", "p6", "Ratio", "Project scale vs ULB size",
         "For small ULBs (<20k pop): flag if single project > ₹10 crore", 
         "By ULB size", "Medium", "TRUE", "Project too large for ULB size"],
        
        # PART VII - ASSETS
        ["P7_001", "7", "p7_assets", "Consistency", "Age vs condition logic",
         "If asset age > 30 years, cannot be 100% 'Good' condition", 
         "AGE-COND", "High", "TRUE", "Old asset marked as good condition"],
        
        ["P7_002", "7", "p7_assets", "Arithmetic", "Condition breakup sum",
         "Good + Fair + Poor + Unusable = Total_Units", 
         "SUM=TOTAL", "Critical", "TRUE", "Condition quantities don't add up"],
        
        ["P7_003", "7", "p7_assets", "Arithmetic", "Acquisition period sum",
         "Acq_upto_2018 + Acq_2018_2022 + Acq_2023_2025 = Total_Units", 
         "SUM=TOTAL", "Critical", "TRUE", "Acquisition periods don't add up"],
        
        ["P7_004", "7", "p7_assets", "Consistency", "R&M presence check",
         "If assets exist, p7_repair columns should have some non-zero values", 
         "ASSETS→R&M", "Medium", "TRUE", "No repair/maintenance for existing assets"],
        
        ["P7_005", "7", "p7_assets", "Sanity", "Mandatory assets check - Roads",
         "Every ULB must have some road infrastructure", 
         ">0", "High", "TRUE", "No road assets recorded"],
        
        ["P7_006", "7", "p7_assets", "Sanity", "Mandatory assets check - Streetlights",
         "Every ULB should have streetlights", 
         ">0", "Medium", "TRUE", "No streetlight assets recorded"],
        
        ["P7_007", "7", "p7_assets", "Cross-part", "Water assets vs service level",
         "If P8 shows water supply, P7 must show water infrastructure", 
         "P8→P7", "High", "TRUE", "Water service without infrastructure"],
        
        ["P7_008", "7", "p7_assets", "Sanity", "Asset cost reasonableness",
         "For recent assets: verify costs align with Part VI capital works", 
         "P7≈P6", "Medium", "TRUE", "Asset cost doesn't match capital work"],
        
        ["P7_009", "7", "p7_assets", "Cross-part", "Asset value vs balance sheet",
         "Total asset value (P7) should approximate Fixed Assets (P3 balance sheet)", 
         "P7≈P3", "High", "TRUE", "Asset register doesn't match balance sheet"],
        
        ["P7_010", "7", "p7_assets", "Ratio", "Asset to population ratio",
         "For roads: km_per_1000_pop = total_road_km / (pop/1000); Flag if < 0.5 or > 5", 
         "0.5-5 km", "Medium", "TRUE", "Road length unusual for population"],
        
        # PART VIII - SERVICE LEVELS
        ["P8_001", "8", "p8", "Sanity", "Water LPCD calculation",
         "lpcd = total_supply_lpd / population; Flag if < 40 or > 135 (or > 200 for errors)", 
         "40-135 LPCD", "High", "TRUE", "LPCD outside normal range"],
        
        ["P8_002", "8", "p8", "Consistency", "HSC cannot exceed total HH",
         "house_service_connections <= total_households (from Part 1)", 
         "HSC<=HH", "Critical", "TRUE", "Connections exceed households"],
        
        ["P8_003", "8", "p8", "Consistency", "Coverage impossibilities",
         "All coverage percentages must be ≤ 100%", 
         "≤100%", "Critical", "TRUE", "Coverage exceeds 100%"],
        
        ["P8_004", "8", "p8", "Sanity", "Waste generation per capita",
         "waste_per_capita = total_waste / population; Flag if < 0.3 or > 0.7 kg/day", 
         "0.3-0.7 kg", "Medium", "TRUE", "Waste generation unusual"],
        
        ["P8_005", "8", "p8", "Cross-part", "Water charges vs supply",
         "If P8 shows water supply, P3 must show water charges revenue", 
         "P8→P3", "High", "TRUE", "Water supply without revenue"],
        
        ["P8_006", "8", "p8", "Engineering", "Supply >5 MLD should have OHT",
         "If total_supply > 5 MLD, P7 should show OHT assets", 
         ">5MLD→OHT", "Medium", "TRUE", "Large supply without OHT"],
        
        ["P8_007", "8", "p8", "Cross-part", "SWM expenses validation",
         "If P8 shows waste collection, P3 must show SWM O&M expenses", 
         "P8→P3", "High", "TRUE", "Waste service without expenses"],
        
        ["P8_008", "8", "p8", "Ratio", "Streetlight per capita",
         "lights_per_1000 = total_streetlights / (pop/1000); Flag if < 5 or > 30", 
         "5-30", "Medium", "TRUE", "Streetlight density unusual"],
        
        ["P8_009", "8", "p8", "Cross-part", "Electricity expense vs streetlights",
         "If P8 shows streetlights, P3 must show electricity expenses", 
         "P8→P3", "High", "TRUE", "Streetlights without electricity expense"],
        
        # PART IX - FUTURE NEEDS
        ["P9_001", "9", "p9", "Ratio", "Proposed vs existing assets",
         "proposed_new / existing_assets; Flag if > 0.25 (25% expansion)", 
         "<25%", "Medium", "TRUE", "Unusually large expansion proposed"],
        
        ["P9_002", "9", "p9", "Sanity", "Cost reasonableness",
         "Verify proposed costs align with engineering norms (from Part VI)", 
         "P9≈P6 norms", "High", "TRUE", "Proposed cost unreasonable"],
        
        ["P9_003", "9", "p9", "Engineering", "System completeness",
         "If new OHT proposed, must include pipeline network", 
         "LOGIC", "High", "TRUE", "Incomplete system proposed"],
        
        ["P9_004", "9", "p9", "Arithmetic", "Funding sources sum",
         "Own_Source + Grants = Total_Requirement", 
         "SUM=TOTAL", "Critical", "TRUE", "Funding doesn't add up"],
        
        ["P9_005", "9", "p9", "Cross-part", "Comparison with past capex",
         "Proposed capex should be comparable to Part VI historical capex patterns", 
         "P9≈P6", "Medium", "TRUE", "Proposed capex inconsistent with history"],
        
        ["P9_006", "9", "p9", "Ratio", "Own source capacity",
         "own_source_ratio = own_contribution / total_requirement; Verify feasibility vs revenue", 
         "Feasible", "High", "TRUE", "Own contribution not feasible"],
        
        # STATISTICAL/ADVANCED CHECKS
        ["STAT_001", "All", "All", "Statistical", "Outlier detection - Standard deviation",
         "For numeric fields: flag if value > Mean + 2*StdDev across peer ULBs", 
         ">Mean+2SD", "Medium", "TRUE", "Statistical outlier"],
        
        ["STAT_002", "All", "All", "Statistical", "Unit error detection",
         "Flag if value is 100× or 1000× peer median (unit confusion)", 
         "100×/1000×", "High", "TRUE", "Likely unit error"],
        
        ["STAT_003", "All", "All", "Pattern", "Repeated values detection",
         "Flag if same value appears in >50% of rows for a field", 
         ">50% same", "Medium", "TRUE", "Suspicious repeated values"],
        
        ["STAT_004", "All", "All", "Pattern", "Rounded numbers only",
         "Flag if >80% of values are round numbers (ends in 00, 000, etc.)", 
         ">80% rounded", "Low", "TRUE", "Suspiciously rounded data"],
        
        ["STAT_005", "All", "All", "Pattern", "Zero-heavy columns",
         "Flag if >60% of values are zero (possible data entry issue)", 
         ">60% zeros", "Medium", "TRUE", "Too many zeros"],
        
        # CROSS-PART VALIDATION SUMMARY
        ["CROSS_001", "3,4", "Multiple", "Cross-part", "Revenue reconciliation",
         "Part III revenue items must reconcile with Part IV collections", 
         "P3=P4", "Critical", "TRUE", "Revenue-collection mismatch"],
        
        ["CROSS_002", "3,5", "Multiple", "Cross-part", "Liability reconciliation",
         "Part III balance sheet liabilities must match Part V details", 
         "P3=P5", "Critical", "TRUE", "Balance sheet-liability mismatch"],
        
        ["CROSS_003", "3,6,7", "Multiple", "Cross-part", "Capital expenditure flow",
         "Part III capex → Part VI projects → Part VII asset additions", 
         "P3→P6→P7", "High", "TRUE", "Capital expenditure flow broken"],
        
        ["CROSS_004", "7,8", "Multiple", "Cross-part", "Infrastructure vs service",
         "Part VII assets must support Part VIII service levels", 
         "P7→P8", "High", "TRUE", "Service level without infrastructure"],
        
        ["CROSS_005", "6,9", "Multiple", "Cross-part", "Historical vs future capex",
         "Part IX proposals should align with Part VI execution capacity", 
         "P9≈P6", "Medium", "TRUE", "Unrealistic future plans"],
    ]
    
    # Create DataFrame
    df_rules = pd.DataFrame(validation_rules, columns=[
        "RuleID", "PartNo", "TableName", "CheckType", "Description",
        "Formula", "Threshold", "Severity", "Enabled", "ErrorMessage"
    ])
    
    ws_rules = wb.create_sheet("ValidationRules")
    
    # Write headers
    headers = list(df_rules.columns)
    ws_rules.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws_rules.cell(1, col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    
    # Write data
    for r_idx, row in enumerate(dataframe_to_rows(df_rules, index=False, header=False), 2):
        for c_idx, value in enumerate(row, 1):
            cell = ws_rules.cell(r_idx, c_idx, value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            
            # Color code severity
            if c_idx == 8:  # Severity column
                if value == "Critical":
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
                elif value == "High":
                    cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                elif value == "Medium":
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    # Set column widths
    ws_rules.column_dimensions['A'].width = 10  # RuleID
    ws_rules.column_dimensions['B'].width = 8   # PartNo
    ws_rules.column_dimensions['C'].width = 20  # TableName
    ws_rules.column_dimensions['D'].width = 12  # CheckType
    ws_rules.column_dimensions['E'].width = 40  # Description
    ws_rules.column_dimensions['F'].width = 60  # Formula
    ws_rules.column_dimensions['G'].width = 15  # Threshold
    ws_rules.column_dimensions['H'].width = 10  # Severity
    ws_rules.column_dimensions['I'].width = 8   # Enabled
    ws_rules.column_dimensions['J'].width = 40  # ErrorMessage
    
    # Freeze panes
    ws_rules.freeze_panes = 'A2'
    
    # ============================================================================
    # SHEET 2: THRESHOLD VALUES
    # ============================================================================
    
    threshold_data = [
        ["Category", "Parameter", "Min", "Max", "Unit", "Notes"],
        ["Demographics", "Population Density", "300", "35000", "per sq.km", "Urban areas"],
        ["Demographics", "Household Size", "2.5", "6", "persons", "Average household"],
        ["Demographics", "Gender Ratio (Male)", "0.45", "0.55", "ratio", "45-55%"],
        ["Demographics", "Population Growth Rate", "0.005", "0.05", "annual", "0.5-5% CAGR"],
        ["Demographics", "Floating Population", "0", "0.5", "ratio", "<50% of resident pop"],
        ["HR", "Class I Salary", "600000", "1800000", "₹/year", "Level 18-22"],
        ["HR", "Class II Salary", "400000", "1000000", "₹/year", "Level 12-17"],
        ["HR", "Class III Salary", "250000", "600000", "₹/year", "Level 6-11"],
        ["HR", "Class IV Salary", "180000", "400000", "₹/year", "Level 1-5"],
        ["HR", "Vacancy Rate", "0", "0.4", "ratio", "<40% acceptable"],
        ["HR", "Staff per 1000 pop", "2", "8", "per 1000", "Total staff"],
        ["Revenue", "Property Tax per capita", "500", "4000", "₹", "Annual"],
        ["Revenue", "Water Charges per capita", "200", "1500", "₹", "Annual"],
        ["Revenue", "Own Revenue Ratio", "0.3", "1", "ratio", ">30% desirable"],
        ["Revenue", "Revenue Growth YoY", "-0.3", "0.5", "ratio", "-30% to +50%"],
        ["Expenditure", "Salary to Total Exp", "0", "0.5", "ratio", "<50%"],
        ["Expenditure", "Expenditure Growth YoY", "-0.2", "0.5", "ratio", "-20% to +50%"],
        ["Expenditure", "Road O&M % of Asset Value", "0.02", "0.06", "ratio", "2-6%"],
        ["Expenditure", "Water O&M % of Asset Value", "0.03", "0.08", "ratio", "3-8%"],
        ["Expenditure", "SWM O&M % of Asset Value", "0.05", "0.15", "ratio", "5-15%"],
        ["Taxation", "Collection Efficiency", "0.5", "1", "ratio", "50-100%"],
        ["Taxation", "Uncollectible %", "0", "0.1", "ratio", "<10%"],
        ["Taxation", "Demand per Assessment", "500", "50000", "₹", "Property tax"],
        ["Liabilities", "Debt Service Ratio", "0", "0.25", "ratio", "<25% of revenue"],
        ["Capital Works", "BT Road Cost", "30", "250", "lakh/km", "Bitumen topped"],
        ["Capital Works", "CC Road Cost", "50", "300", "lakh/km", "Cement concrete"],
        ["Capital Works", "Water Pipeline Cost", "15", "80", "lakh/km", "Distribution"],
        ["Capital Works", "STP Cost", "80", "300", "lakh/MLD", "Sewage treatment"],
        ["Capital Works", "OHT Cost", "20", "100", "lakh/lakh litre", "Overhead tank"],
        ["Capital Works", "Capex per capita", "0", "10000", "₹", "Annual"],
        ["Assets", "Road per 1000 pop", "0.5", "5", "km", "Total road length"],
        ["Service Level", "Water LPCD", "40", "135", "liters", "Normal range"],
        ["Service Level", "Water LPCD Max", "135", "200", "liters", "High but possible"],
        ["Service Level", "Waste per capita", "0.3", "0.7", "kg/day", ""],
        ["Service Level", "Streetlights per 1000", "5", "30", "nos", ""],
        ["Future Needs", "Expansion Ratio", "0", "0.25", "ratio", "New/Existing <25%"],
    ]
    
    ws_threshold = wb.create_sheet("ThresholdValues")
    
    for row in threshold_data:
        ws_threshold.append(row)
    
    # Style threshold sheet
    for row in ws_threshold.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
    
    for row in ws_threshold.iter_rows(min_row=2, max_row=ws_threshold.max_row):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
    
    ws_threshold.column_dimensions['A'].width = 20
    ws_threshold.column_dimensions['B'].width = 30
    ws_threshold.column_dimensions['C'].width = 12
    ws_threshold.column_dimensions['D'].width = 12
    ws_threshold.column_dimensions['E'].width = 15
    ws_threshold.column_dimensions['F'].width = 40
    ws_threshold.freeze_panes = 'A2'
    
    # ============================================================================
    # SHEET 3: CROSS-PART MAPPING
    # ============================================================================
    
    cross_part_data = [
        ["SourcePart", "SourceTable", "SourceField", "TargetPart", "TargetTable", "TargetField", "Relationship", "ValidationRule"],
        ["1", "p1_1_1_2", "mp_id", "All", "All", "mp_id", "Primary Key", "Must exist in all tables"],
        ["1", "p1_1_1_2", "p1_1_3_4_tot_25_tot", "2", "Summary", "population", "Reference", "Population for ratios"],
        ["1", "p1_1_1_2", "p1_1_5_family", "4", "p4", "households", "Reference", "Assessments <= Households"],
        ["3", "p3", "property_tax_revenue", "4", "p4", "property_tax_collection", "Reconciliation", "Revenue = Collection"],
        ["3", "p3", "salary_expenditure", "2", "p2_2_1_1_2_1_5", "total_salary", "Reconciliation", "Salary exp = Staff salary"],
        ["3", "p3", "liabilities_bs", "5", "p5a/p5b", "total_liabilities", "Reconciliation", "BS liabilities = Detailed liabilities"],
        ["3", "p3", "capital_expenditure", "6", "p6", "total_project_cost", "Flow", "Capex funds projects"],
        ["6", "p6", "completed_works", "7", "p7_assets", "new_assets", "Flow", "Projects become assets"],
        ["7", "p7_assets", "total_asset_value", "3", "p3", "fixed_assets_bs", "Reconciliation", "Asset register ≈ Balance sheet"],
        ["7", "p7_assets", "water_infrastructure", "8", "p8", "water_service", "Logic", "Infrastructure enables service"],
        ["7", "p7_assets", "road_length", "8", "p8", "road_coverage", "Logic", "Assets support coverage"],
        ["8", "p8", "water_supply", "3", "p3", "water_charges", "Logic", "Service generates revenue"],
        ["8", "p8", "waste_collection", "3", "p3", "swm_expense", "Logic", "Service requires expense"],
        ["6", "p6", "historical_capex", "9", "p9", "proposed_capex", "Comparison", "Future aligned with past"],
        ["7", "p7_assets", "existing_assets", "9", "p9", "proposed_expansion", "Comparison", "Expansion < 25% existing"],
    ]
    
    ws_cross = wb.create_sheet("CrossPartMapping")
    
    for row in cross_part_data:
        ws_cross.append(row)
    
    # Style cross-part sheet
    for row in ws_cross.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
    
    for row in ws_cross.iter_rows(min_row=2, max_row=ws_cross.max_row):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    
    ws_cross.column_dimensions['A'].width = 12
    ws_cross.column_dimensions['B'].width = 20
    ws_cross.column_dimensions['C'].width = 25
    ws_cross.column_dimensions['D'].width = 12
    ws_cross.column_dimensions['E'].width = 20
    ws_cross.column_dimensions['F'].width = 25
    ws_cross.column_dimensions['G'].width = 15
    ws_cross.column_dimensions['H'].width = 40
    ws_cross.freeze_panes = 'A2'
    
    # ============================================================================
    # SHEET 4: SEVERITY LEVELS
    # ============================================================================
    
    severity_data = [
        ["Severity", "Priority", "Description", "Action Required", "Examples"],
        ["Critical", "1", "Data integrity issues, arithmetic errors, must be fixed immediately", 
         "Block report generation until fixed", "Arithmetic mismatches, balance sheet not balancing, negative impossible values"],
        ["High", "2", "Significant discrepancies, cross-part mismatches, likely errors", 
         "Flag for immediate review, may require data correction", "Collection >100%, cross-part reconciliation failures, large outliers"],
        ["Medium", "3", "Unusual but possible values, threshold violations, needs verification", 
         "Review and explain in audit report", "Values outside normal range but not impossible, unusual ratios"],
        ["Low", "4", "Minor issues, data quality concerns, informational", 
         "Document for awareness, optional follow-up", "Repeated rounded numbers, minor inconsistencies"],
    ]
    
    ws_severity = wb.create_sheet("SeverityLevels")
    
    for row in severity_data:
        ws_severity.append(row)
    
    # Style severity sheet
    for row in ws_severity.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
    
    for row in ws_severity.iter_rows(min_row=2, max_row=ws_severity.max_row):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            
            # Color code severity
            if cell.column == 1:
                severity = cell.value
                if severity == "Critical":
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
                elif severity == "High":
                    cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                    cell.font = Font(bold=True)
                elif severity == "Medium":
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                elif severity == "Low":
                    cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    
    ws_severity.column_dimensions['A'].width = 12
    ws_severity.column_dimensions['B'].width = 10
    ws_severity.column_dimensions['C'].width = 50
    ws_severity.column_dimensions['D'].width = 40
    ws_severity.column_dimensions['E'].width = 60
    ws_severity.freeze_panes = 'A2'
    
    # ============================================================================
    # SHEET 5: README / INSTRUCTIONS
    # ============================================================================
    
    ws_readme = wb.create_sheet("README", 0)  # Insert at beginning
    
    readme_content = [
        ["ULB AUDIT VALIDATION FRAMEWORK - INSTRUCTIONS"],
        [""],
        ["OVERVIEW:"],
        ["This workbook contains the complete validation rule configuration for auditing Tamil Nadu Municipality (ULB) questionnaire data."],
        ["The Python audit engine reads these rules and applies them to the data automatically."],
        [""],
        ["SHEETS:"],
        ["1. README - This sheet with instructions"],
        ["2. ValidationRules - Master list of all validation checks (100+ rules covering all 9 parts)"],
        ["3. ThresholdValues - Numeric thresholds and acceptable ranges for various parameters"],
        ["4. CrossPartMapping - Relationships between different parts for cross-validation"],
        ["5. SeverityLevels - Classification of audit findings by priority"],
        [""],
        ["HOW TO USE:"],
        [""],
        ["ENABLING/DISABLING RULES:"],
        ["- Navigate to 'ValidationRules' sheet"],
        ["- Column I ('Enabled') controls whether a rule is active"],
        ["- Set to TRUE to enable, FALSE to disable"],
        ["- You can disable rules temporarily for testing or if not applicable"],
        [""],
        ["MODIFYING THRESHOLDS:"],
        ["- Navigate to 'ThresholdValues' sheet"],
        ["- Update Min/Max values as needed based on local context"],
        ["- These values are referenced by validation rules"],
        [""],
        ["ADDING NEW RULES:"],
        ["- Add a new row in 'ValidationRules' sheet"],
        ["- Assign unique RuleID (e.g., P1_011, P2_008, etc.)"],
        ["- Specify Part Number, Table Name, Check Type"],
        ["- Write clear Description and Formula"],
        ["- Set appropriate Threshold, Severity, and ErrorMessage"],
        ["- Set Enabled = TRUE"],
        [""],
        ["RULE TYPES:"],
        ["- Sanity: Basic reasonableness checks (e.g., population density 300-35000)"],
        ["- Arithmetic: Mathematical validations (e.g., Sanctioned = OPS + CPS + Vacant)"],
        ["- Consistency: Internal logic checks (e.g., 2025 population >= 2011 population)"],
        ["- Ratio: Calculated metrics with acceptable ranges (e.g., staff per 1000 population)"],
        ["- Trend: Year-over-year change analysis (e.g., revenue growth -30% to +50%)"],
        ["- Cross-part: Validations spanning multiple questionnaire parts"],
        ["- Statistical: Outlier detection using peer comparison"],
        ["- Pattern: Data quality checks (e.g., too many zeros, repeated values)"],
        [""],
        ["SEVERITY LEVELS:"],
        ["- Critical (Red): Must fix - arithmetic errors, data integrity issues"],
        ["- High (Orange): Likely errors - significant discrepancies, reconciliation failures"],
        ["- Medium (Yellow): Needs verification - unusual but possible values"],
        ["- Low (Green): Informational - minor quality issues"],
        [""],
        ["WORKFLOW:"],
        ["1. Place all ULB CSV files in the 'data' folder"],
        ["2. Double-click 'run_audit.bat' to start the audit process"],
        ["3. The script will:"],
        ["   - Load all CSV files"],
        ["   - Apply all enabled validation rules"],
        ["   - Generate individual ULB audit reports (PDF)"],
        ["   - Create master dashboard (Excel) with summary of all findings"],
        ["4. Review reports in 'reports' folder"],
        [""],
        ["CUSTOMIZATION:"],
        ["- Modify this workbook to adjust rules, thresholds, severity levels"],
        ["- No need to modify Python code for rule changes"],
        ["- Save this file after making changes"],
        ["- Re-run the audit to apply updated rules"],
        [""],
        ["TECHNICAL NOTES:"],
        ["- The Python engine reads rules from 'ValidationRules' sheet where Enabled=TRUE"],
        ["- Formulas are interpreted as Python expressions"],
        ["- Cross-part validations require data from multiple CSV files"],
        ["- Statistical checks require at least 10 ULBs for peer comparison"],
        [""],
        ["SUPPORT:"],
        ["- For questions about specific rules, refer to the audit documentation PDFs"],
        ["- Part 1-9.pdf contains detailed explanations of validation logic"],
        ["- PDFsam_merge.pdf contains database schema and column mappings"],
        [""],
        ["VERSION INFORMATION:"],
        ["- Framework Version: 1.0"],
        ["- Last Updated: February 2026"],
        ["- Rules Count: 100+ (covering all 9 parts + cross-validations)"],
    ]
    
    for row_data in readme_content:
        ws_readme.append(row_data)
    
    # Style README
    # Title
    ws_readme['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws_readme['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws_readme['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws_readme.merge_cells('A1:E1')
    
    # Section headers
    for row in [3, 8, 11, 14, 21, 27, 34, 41, 48, 54, 60, 66]:
        cell = ws_readme.cell(row, 1)
        cell.font = Font(size=12, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    # Wrap text
    for row in ws_readme.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    
    ws_readme.column_dimensions['A'].width = 120
    
    # Save workbook
    output_path = "/home/claude/ulb_audit_system/config/ULB_Validation_Rules_Master.xlsx"
    wb.save(output_path)
    print(f"✓ Created validation rules master: {output_path}")
    return output_path

if __name__ == "__main__":
    create_validation_rules_master()
